import json
from numpy import double
from openpyxl import load_workbook


class HBLibCustomiser:
    @staticmethod
    def _getValue(value, valueType: str):
        if valueType == "字符串":
            return str(value)
        elif valueType == "整数":
            return int(value)
        elif valueType == "小数":
            return float(value)
        else:
            return None

    @staticmethod
    def loadExcel(excelPath: str) -> dict:
        # 读取数据
        wb = load_workbook(excelPath)
        ws = wb["opaque_material"]
        #for colIndex in range(2,ws.max_column+1):

        materialsDict = dict()
        for i in range(ws.max_row-2):
            materialParams = dict()
            rowIndex = i+3
            name = ws.cell(rowIndex, 1).value
            for colIndex in range(2, ws.max_column):
                value = ws.cell(rowIndex, colIndex).value
                valueType = str(ws.cell(1, colIndex).value)
                value = HBLibCustomiser._getValue(value, valueType)
                valueName = str(ws.cell(2, colIndex).value)
                if value is None or value == "" or (value is str and value.isspace()) or valueName is None or valueName == "" or valueName.isspace():
                    continue
                
                materialParams[valueName] = value

            materialsDict[name] = materialParams
        
        return materialsDict


    def createJson(JsonPath: str, data: dict) -> None:
        # 写入 JSON 数据
        with open(JsonPath, 'w') as f:
            json.dump(data, f,indent=2)
